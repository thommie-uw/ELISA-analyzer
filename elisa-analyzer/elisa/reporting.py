"""Excel report builder (multi-sheet .xlsx with the curve image embedded)."""
from __future__ import annotations

import io
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analysis import AnalysisResult, display_wells, plate_matrix
from .layout import parse_code
from .plate import N_COLS, N_ROWS, ROWS

HEADER_FILL = PatternFill("solid", fgColor="1B3A56")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1B3A56")
SUB_FONT = Font(bold=True, size=11, color="1B3A56")
FLAG_FILL = PatternFill("solid", fgColor="FCE4E2")
WARN_FILL = PatternFill("solid", fgColor="FFF4DA")
OK_FILL = PatternFill("solid", fgColor="EAF4EC")
THIN = Side(style="thin", color="D6DBE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROLE_FILL = {
    "standard": PatternFill("solid", fgColor="DCE9F4"),
    "blank": PatternFill("solid", fgColor="EDEFF2"),
    "sample": PatternFill("solid", fgColor="FDF0DC"),
    "control": PatternFill("solid", fgColor="E7E2F5"),
    "nsb": PatternFill("solid", fgColor="F6E6E6"),
    "empty": PatternFill("solid", fgColor="FFFFFF"),
}


def _write_frame(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1,
                 number_format: str = "0.000", freeze: bool = True) -> int:
    """Write a DataFrame with a styled header. Returns the next free row."""
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    flag_cols = {
        j for j, c in enumerate(df.columns)
        if str(c).lower().endswith("flag") or str(c).lower() == "flags"
    }
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            val = row[col]
            if isinstance(val, (np.floating, np.integer)):
                val = val.item()
            if isinstance(val, float) and not np.isfinite(val):
                val = None
            if isinstance(val, (np.bool_, bool)):
                val = bool(val)
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.border = BORDER
            if isinstance(val, float):
                cell.number_format = "0.0" if "%" in str(col) else number_format
            if j in flag_cols and isinstance(val, str):
                if val in ("OK", "In Range"):
                    cell.fill = OK_FILL
                elif val != "n/a":
                    cell.fill = FLAG_FILL

    for j, col in enumerate(df.columns):
        width = max(len(str(col)) + 2, 10)
        sample = df[col].astype(str).head(60)
        if len(sample):
            width = max(width, min(38, int(sample.str.len().max()) + 3))
        ws.column_dimensions[get_column_letter(start_col + j)].width = width

    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    return start_row + len(df) + 2


def _write_plate_grid(ws, matrix, title: str, start_row: int, layout=None,
                      number_format: str = "0.000") -> int:
    ws.cell(row=start_row, column=1, value=title).font = SUB_FONT
    hdr = start_row + 1
    for c in range(N_COLS):
        cell = ws.cell(row=hdr, column=2 + c, value=c + 1)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r in range(N_ROWS):
        rc = ws.cell(row=hdr + 1 + r, column=1, value=ROWS[r])
        rc.fill = HEADER_FILL
        rc.font = HEADER_FONT
        rc.alignment = Alignment(horizontal="center")
        for c in range(N_COLS):
            v = matrix[r][c] if isinstance(matrix, list) else matrix[r, c]
            if isinstance(v, (np.floating, np.integer)):
                v = v.item()
            if isinstance(v, float) and not np.isfinite(v):
                v = None
            cell = ws.cell(row=hdr + 1 + r, column=2 + c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if isinstance(v, float):
                cell.number_format = number_format
            if layout is not None:
                cell.fill = ROLE_FILL.get(parse_code(layout[r][c] if isinstance(layout, list)
                                                     else layout[r, c]).role, ROLE_FILL["empty"])
    for c in range(N_COLS + 1):
        ws.column_dimensions[get_column_letter(c + 1)].width = 10
    return hdr + N_ROWS + 3


def build_report(
    result: AnalysisResult,
    curve_png: bytes | None = None,
    residual_png: bytes | None = None,
    meta: dict | None = None,
    layout=None,
) -> bytes:
    meta = meta or {}
    units = result.units
    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "ELISA Analysis Report"
    ws["A1"].font = TITLE_FONT
    row = 3
    info = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source File", meta.get("source_file", "")),
        ("Plate / Read", meta.get("plate_name", "")),
        ("Assay Name", meta.get("assay_name", "")),
        ("Analyst", meta.get("analyst", "")),
        ("Concentration Units", units),
        ("Blank Subtraction", "Yes" if result.options.subtract_blank else "No"),
        ("Curve Fitted On", result.options.fit_on),
    ]
    for k, v in info:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=v)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Quality Summary").font = SUB_FONT
    row = _write_frame(ws, result.qc_summary(), start_row=row + 1, freeze=False)

    ws.cell(row=row, column=1, value="Fit Parameters").font = SUB_FONT
    params = pd.DataFrame(
        [
            {"Parameter": name, "Value": val,
             "Std. Error": result.fit.stderr.get(key, np.nan)}
            for (name, val), key in zip(result.fit.summary_rows(), result.fit.param_order)
        ]
    )
    row = _write_frame(ws, params, start_row=row + 1, number_format="0.00000", freeze=False)
    ws.cell(row=row, column=1, value="Equation").font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value=result.fit.equation())
    row += 2

    if result.warnings:
        ws.cell(row=row, column=1, value="Notes").font = SUB_FONT
        row += 1
        for w in result.warnings:
            c = ws.cell(row=row, column=1, value=w)
            c.fill = WARN_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16

    if curve_png:
        img = XLImage(io.BytesIO(curve_png))
        img.width, img.height = 620, 410
        ws.add_image(img, f"E3")
    if residual_png:
        img2 = XLImage(io.BytesIO(residual_png))
        img2.width, img2.height = 620, 215
        ws.add_image(img2, f"E28")

    # ---------------- results ----------------
    ws2 = wb.create_sheet("Sample Results")
    _write_frame(ws2, result.samples)

    if len(result.controls):
        wsc = wb.create_sheet("Controls")
        _write_frame(wsc, result.controls)

    ws3 = wb.create_sheet("Standards")
    _write_frame(ws3, result.standards)

    ws4 = wb.create_sheet("Well Data")
    _write_frame(ws4, display_wells(result.wells, units))

    # ---------------- plate maps ----------------
    ws5 = wb.create_sheet("Plate Maps")
    r = 1
    lay = None
    if layout is not None:
        lay = [[str(layout[i, j]) for j in range(N_COLS)] for i in range(N_ROWS)]
        r = _write_plate_grid(ws5, lay, "Layout", r, layout=layout, number_format="@")
    r = _write_plate_grid(ws5, plate_matrix(result.wells, "OD Raw"),
                          "Raw Absorbance", r, layout=layout)
    r = _write_plate_grid(ws5, plate_matrix(result.wells, "OD Corrected"),
                          "Blank-Corrected Absorbance", r, layout=layout)
    r = _write_plate_grid(ws5, plate_matrix(result.wells, "Conc"),
                          f"Back-Calculated Concentration ({units})", r,
                          layout=layout, number_format="0.00")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def results_csv(result: AnalysisResult) -> bytes:
    return result.samples.to_csv(index=False).encode("utf-8")
