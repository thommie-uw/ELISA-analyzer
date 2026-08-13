"""Unit tests for parsing, curve fitting, layout handling and analysis."""
from __future__ import annotations

import json
import math
import os
import sys

import numpy as np
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from elisa.analysis import AnalysisOptions, analyse  # noqa: E402
from elisa.fitting import (  # noqa: E402
    fit_standard_curve,
    five_pl,
    five_pl_inverse,
    four_pl,
    four_pl_inverse,
)
from elisa.layout import (  # noqa: E402
    autofill_samples,
    blank_layout,
    from_template,
    parse_code,
    preset_layout,
    summarise,
    to_template,
    validate,
)
from elisa.parsing import read_plates, to_number  # noqa: E402
from elisa.reporting import build_report  # noqa: E402

SAMPLES = os.path.join(ROOT, "sample_data")
STD_CONC = [1000.0, 500.0, 250.0, 125.0, 62.5, 31.25, 15.625]
TRUE = dict(a=0.048, b=1.15, c=42.0, d=2.72)


def approx(a, b, tol=1e-6):
    return abs(a - b) <= tol * max(1.0, abs(b))


# --------------------------------------------------------------------------
# fitting
# --------------------------------------------------------------------------
def test_4pl_recovers_known_parameters():
    x = np.repeat(np.array(STD_CONC + [0.0]), 3)
    y = four_pl(x, **TRUE)
    fit = fit_standard_curve(x, y, "4PL")
    for k, v in TRUE.items():
        assert approx(fit.params[k], v, 1e-3), f"{k}: {fit.params[k]} vs {v}"
    assert fit.r_squared > 0.9999


def test_5pl_recovers_known_parameters():
    true5 = dict(a=0.05, b=1.3, c=60.0, d=2.9, g=0.7)
    x = np.repeat(np.array(STD_CONC + [0.0]), 3)
    y = five_pl(x, **true5)
    fit = fit_standard_curve(x, y, "5PL")
    pred = fit.predict(x)
    assert np.allclose(pred, y, atol=1e-4), "5PL fit does not reproduce the data"
    assert fit.r_squared > 0.9999


def test_inverse_is_exact():
    x = np.array([1.0, 10.0, 100.0, 1000.0])
    y4 = four_pl(x, **TRUE)
    assert np.allclose(four_pl_inverse(y4, **TRUE), x, rtol=1e-8)
    true5 = dict(a=0.05, b=1.3, c=60.0, d=2.9, g=0.7)
    y5 = five_pl(x, **true5)
    assert np.allclose(five_pl_inverse(y5, **true5), x, rtol=1e-8)


def test_inverse_returns_nan_outside_asymptotes():
    out = four_pl_inverse(np.array([5.0, -1.0]), **TRUE)
    assert np.isnan(out).all(), "values beyond the asymptotes must not extrapolate"


def test_decreasing_competitive_curve():
    true = dict(a=2.6, b=1.1, c=25.0, d=0.07)
    x = np.repeat(np.array(STD_CONC), 2)
    y = four_pl(x, **true)
    fit = fit_standard_curve(x, y, "4PL")
    assert fit.r_squared > 0.9999
    assert np.allclose(fit.inverse(fit.predict(x)), x, rtol=1e-4)


def test_weighting_helps_the_low_end():
    """With proportional error, 1/Y^2 weighting should track the low standards better.

    Averaged over several simulated plates -- on any single plate the noise can
    go either way.
    """
    x = np.repeat(np.array(STD_CONC), 3)
    low = x <= 62.5
    target = four_pl(x[low], **TRUE)
    err_plain, err_weighted = [], []
    for seed in range(12):
        rng = np.random.default_rng(seed)
        y = four_pl(x, **TRUE) * (1 + rng.normal(0, 0.03, x.size))
        plain = fit_standard_curve(x, y, "4PL", "None")
        weighted = fit_standard_curve(x, y, "4PL", "1/Y²")
        err_plain.append(np.abs((plain.predict(x[low]) - target) / target).mean())
        err_weighted.append(np.abs((weighted.predict(x[low]) - target) / target).mean())
    assert np.mean(err_weighted) < np.mean(err_plain), (
        f"weighted {np.mean(err_weighted):.5f} vs plain {np.mean(err_plain):.5f}"
    )


def test_too_few_points_raises():
    try:
        fit_standard_curve([1.0, 2.0], [0.1, 0.2], "4PL")
    except ValueError:
        return
    raise AssertionError("expected a ValueError for an under-determined fit")


# --------------------------------------------------------------------------
# parsing
# --------------------------------------------------------------------------
def test_all_export_formats_agree():
    files = [
        "plate_generic.csv",
        "plate_softmax_pro.txt",
        "plate_gen5.xlsx",
        "plate_well_list.csv",
    ]
    mats = []
    for name in files:
        with open(os.path.join(SAMPLES, name), "rb") as fh:
            reads = read_plates(fh.read(), name)
        assert len(reads) == 1, f"{name}: expected 1 block, got {len(reads)}"
        assert reads[0].values.shape == (8, 12)
        mats.append(reads[0].values)
    for name, m in zip(files[1:], mats[1:]):
        assert np.allclose(mats[0], m, equal_nan=True), f"{name} differs from the CSV"


def test_number_parsing_edge_cases():
    assert to_number("0.512")[0] == 0.512
    assert to_number("0,512")[0] == 0.512          # European decimal comma
    assert to_number("1.2e-1")[0] == 0.12
    assert math.isnan(to_number("")[0])
    val, over = to_number("OVRFLW")
    assert math.isnan(val) and over is True
    assert math.isnan(to_number("Temperature")[0])


def test_two_wavelengths_side_by_side():
    rows = ["\t".join([""] + [str(c) for c in range(1, 13)] * 2)]
    rng = np.random.default_rng(1)
    a = np.round(rng.uniform(0.1, 2.5, (8, 12)), 4)
    b = np.round(rng.uniform(0.02, 0.09, (8, 12)), 4)
    for i, r in enumerate("ABCDEFGH"):
        rows.append("\t".join([r] + [f"{v}" for v in a[i]] + [f"{v}" for v in b[i]]))
    reads = read_plates("\n".join(rows).encode(), "dual.txt")
    assert len(reads) == 2, f"expected 2 blocks, got {len(reads)}"
    found = [r.values for r in reads]
    assert any(np.allclose(f, a) for f in found)
    assert any(np.allclose(f, b) for f in found)


def test_missing_wells_are_nan_not_errors():
    lines = [",".join([""] + [str(c) for c in range(1, 13)])]
    for i, r in enumerate("ABCDEFGH"):
        vals = ["" if (i == 7 and j > 9) else f"{0.1 + 0.01 * j:.3f}" for j in range(12)]
        lines.append(",".join([r] + vals))
    reads = read_plates("\n".join(lines).encode(), "gaps.csv")
    assert len(reads) == 1
    assert np.isnan(reads[0].values[7, 10:]).all()
    assert reads[0].n_missing == 2


# --------------------------------------------------------------------------
# layout
# --------------------------------------------------------------------------
def test_layout_code_parsing():
    cases = {
        "S1": ("standard", 1), "s12": ("standard", 12), "Std 3": ("standard", 3),
        "B": ("blank", 0), "blank": ("blank", 0), "NSB": ("nsb", 0),
        "QC1": ("control", 0), "ctrl-high": ("control", 0),
        "": ("empty", 0), "-": ("empty", 0), "x": ("empty", 0),
        "Patient 12": ("sample", 0), "S13": ("sample", 0),
    }
    for code, (role, level) in cases.items():
        got = parse_code(code)
        assert got.role == role, f"{code!r} -> {got.role}, expected {role}"
        assert got.level == level, f"{code!r} level {got.level}, expected {level}"


def test_preset_and_autofill():
    lay = preset_layout("Standards Col 1-2, 7 Levels + Blank (Duplicate)")
    info = summarise(lay)
    assert info.standard_levels == list(range(1, 8))
    assert info.n_blank == 2
    filled = autofill_samples(lay, replicates=2, order="column", prefix="Smp")
    info2 = summarise(filled)
    assert len(info2.sample_labels) == 40, len(info2.sample_labels)
    assert all(v == 2 for k, v in info2.replicate_counts.items() if k.startswith("Smp"))
    assert summarise(lay).n_empty == 80, "autofill must not mutate the input layout"


def test_template_round_trip():
    lay = preset_layout("Standards Col 1-2, 8 Levels (Duplicate)")
    lay = autofill_samples(lay, 3, "column", "P")
    conc = {f"S{i + 1}": c for i, c in enumerate(STD_CONC + [7.8])}
    text = to_template(lay, conc, "ng/mL", {"P01": 5.0}, "run A")
    back = from_template(text)
    assert (back["layout"] == lay).all()
    assert back["standards"] == conc
    assert back["dilutions"] == {"P01": 5.0}
    assert back["units"] == "ng/mL"
    json.loads(text)  # must be valid JSON


def test_validate_flags_missing_pieces():
    lay = blank_layout()
    problems = validate(lay, {})
    assert any("No standard wells" in p for p in problems)
    lay2 = preset_layout("Standards Col 1-2, 8 Levels (Duplicate)")
    problems2 = validate(lay2, {"S1": 100.0})
    assert any("without a concentration" in p for p in problems2)


# --------------------------------------------------------------------------
# analysis
# --------------------------------------------------------------------------
def _demo():
    with open(os.path.join(SAMPLES, "plate_softmax_pro.txt"), "rb") as fh:
        plate = read_plates(fh.read(), "plate_softmax_pro.txt")[0]
    with open(os.path.join(SAMPLES, "demo_layout.json"), encoding="utf-8") as fh:
        tpl = from_template(fh.read())
    return plate, tpl


def test_end_to_end_accuracy():
    plate, tpl = _demo()
    truth = pd.read_csv(os.path.join(SAMPLES, "_expected_concentrations.csv"))
    truth = truth.set_index("Sample")["True concentration (pg/mL)"]
    for model in ("4PL", "5PL"):
        res = analyse(plate.values, tpl["layout"], tpl["standards"],
                      AnalysisOptions(model=model, units="pg/mL"))
        assert res.fit.r_squared > 0.999
        idx = res.samples.set_index("Sample")
        got = idx["Mean Conc (pg/mL)"]
        # Sample11 carries a deliberate 35% pipetting error and is flagged as such;
        # accuracy is only meaningful for in-range samples that passed QC.
        ok = (idx["Range Flag"] == "In Range") & (idx["CV Flag"] == "OK")
        comp = pd.DataFrame({"t": truth, "c": got, "ok": ok}).dropna()
        comp = comp[comp["ok"]]
        err = ((comp["c"] / comp["t"] - 1) * 100).abs()
        assert err.median() < 5, f"{model}: median error {err.median():.2f}%"
        assert err.max() < 25, f"{model}: worst error {err.max():.2f}%"


def test_blank_subtraction_and_cv():
    plate, tpl = _demo()
    res = analyse(plate.values, tpl["layout"], tpl["standards"],
                  AnalysisOptions(units="pg/mL", subtract_blank=True))
    blank_wells = res.wells[res.wells["Role"] == "blank"]
    assert len(blank_wells) == 2
    assert approx(res.blank_mean, blank_wells["OD Raw"].mean(), 1e-9)
    assert abs(res.wells["OD Corrected"].mean()
               - (res.wells["OD Raw"].mean() - res.blank_mean)) < 1e-9

    raw = analyse(plate.values, tpl["layout"], tpl["standards"],
                  AnalysisOptions(units="pg/mL", subtract_blank=False))
    assert approx(raw.wells["OD Corrected"].sum(), raw.wells["OD Raw"].sum(), 1e-9)

    # %CV against a hand calculation for one sample
    row = res.samples.iloc[0]
    wells = [w.strip() for w in row["Wells"].split(",")]
    ods = res.wells.set_index("Well").loc[wells, "OD Corrected"].to_numpy()
    expected = ods.std(ddof=1) / ods.mean() * 100
    assert approx(row["%CV OD"], expected, 1e-9)


def test_high_cv_is_flagged():
    plate, tpl = _demo()
    res = analyse(plate.values, tpl["layout"], tpl["standards"],
                  AnalysisOptions(units="pg/mL", cv_threshold=20.0))
    flagged = res.samples[res.samples["CV Flag"] == "High CV"]["Sample"].tolist()
    assert "Sample11" in flagged, f"expected the noisy pair to be flagged, got {flagged}"


def test_out_of_range_flags():
    plate, tpl = _demo()
    res = analyse(plate.values, tpl["layout"], tpl["standards"], AnalysisOptions(units="pg/mL"))
    flags = res.samples.set_index("Sample")["Range Flag"]
    assert flags["Sample03"] in ("> ULOQ", "Above Curve (> ULOQ)")
    assert flags["Sample07"] in ("< LLOQ", "Below Curve (< LLOQ)")
    assert res.lloq == min(c for c in tpl["standards"].values() if c > 0)
    assert res.uloq == max(tpl["standards"].values())


def test_dilution_factor_is_applied():
    plate, tpl = _demo()
    base = analyse(plate.values, tpl["layout"], tpl["standards"], AnalysisOptions(units="pg/mL"))
    dil = analyse(plate.values, tpl["layout"], tpl["standards"],
                  AnalysisOptions(units="pg/mL", dilutions={"Sample01": 10.0}))
    b = base.samples.set_index("Sample")
    d = dil.samples.set_index("Sample")
    assert approx(d.loc["Sample01", "Final Conc (pg/mL)"],
                  b.loc["Sample01", "Final Conc (pg/mL)"] * 10, 1e-9)
    assert approx(d.loc["Sample02", "Final Conc (pg/mL)"],
                  b.loc["Sample02", "Final Conc (pg/mL)"], 1e-9)


def test_excluding_a_well_changes_the_group():
    plate, tpl = _demo()
    base = analyse(plate.values, tpl["layout"], tpl["standards"], AnalysisOptions(units="pg/mL"))
    target = base.samples.iloc[5]
    well = target["Wells"].split(",")[0].strip()
    excl = analyse(plate.values, tpl["layout"], tpl["standards"],
                   AnalysisOptions(units="pg/mL", excluded_wells={well}))
    row = excl.samples[excl.samples["Sample"] == target["Sample"]].iloc[0]
    assert row["N"] == target["N"] - 1
    assert well not in row["Wells"]


def test_excluding_a_standard_level_shrinks_the_fit():
    plate, tpl = _demo()
    base = analyse(plate.values, tpl["layout"], tpl["standards"], AnalysisOptions(units="pg/mL"))
    cut = analyse(plate.values, tpl["layout"], tpl["standards"],
                  AnalysisOptions(units="pg/mL", excluded_standard_levels={1}))
    assert cut.fit.n_points == base.fit.n_points - 2
    assert cut.uloq < base.uloq
    assert not cut.standards.set_index("Standard").loc["S1", "Used in Fit"]


def test_standard_recovery_within_limits():
    plate, tpl = _demo()
    res = analyse(plate.values, tpl["layout"], tpl["standards"], AnalysisOptions(units="pg/mL"))
    rec = res.standards["Recovery %"].dropna()
    assert (rec.between(80, 120)).all(), res.standards[["Standard", "Recovery %"]]
    assert (res.standards["Recovery Flag"] == "OK").all()


def test_report_is_a_valid_workbook():
    from openpyxl import load_workbook
    import io

    plate, tpl = _demo()
    res = analyse(plate.values, tpl["layout"], tpl["standards"], AnalysisOptions(units="pg/mL"))
    blob = build_report(res, meta={"source_file": "x.txt"}, layout=tpl["layout"])
    wb = load_workbook(io.BytesIO(blob))
    for sheet in ("Summary", "Sample Results", "Standards", "Well Data", "Plate Maps"):
        assert sheet in wb.sheetnames, f"missing sheet {sheet}"
    assert wb["Sample Results"].max_row == len(res.samples) + 1


if __name__ == "__main__":
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failed = 0
    for fn in tests:
        try:
            fn()
            print(f"  ok    {fn.__name__}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"  FAIL  {fn.__name__}: {exc}")
    print(f"\n{len(tests) - failed}/{len(tests)} passed")
    sys.exit(1 if failed else 0)
